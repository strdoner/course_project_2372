from celery import shared_task
from .models import Chart
from .serializers import ChartModelSerializer

import openpyxl
from dataclasses import dataclass
from typing import List, Union, IO
from enum import Enum
import json

@shared_task
def process_uploaded_file(file_path, user_id):
    print(file_path)
    data = parse_excel(file_path)
    print(data)
    for sheet in data:
        sheet_data = {
                'user':user_id,
                'title':sheet.get('title'),
                'min_x':1,
                'min_y':1,
                'max_x':1,
                'max_y':1,
                'keys':sheet.get('keys')
        }
        serializer = ChartModelSerializer(data=sheet_data)
        if serializer.is_valid():
            serializer.save()
            print(serializer.data)


    return {"status": "success", "result": data}


class ErrorType(Enum):
    MISSING_HEADERS = "Missing headers for key and value names."
    NO_FILE_PROVIDED = "No file provided."
    FILE_PROCESSING_ERROR = "Error processing the file."

@dataclass
class CoordinateMapping:
    sheet_name: str
    x_name: str
    y_name: str
    mapping: List[tuple[Union[int, float], Union[int, float]]]
    error: str = None

    def to_json(self) -> str:
        
        self.mapping = sorted(self.mapping, key=lambda x: x[0])

        x_list, y_list = zip(*self.mapping) if self.mapping else ([], [])
        if len(x_list) != len(y_list) or len(x_list) == 0:
            return None 
        format = {
            "title" : self.sheet_name,
            "keys" : {
                "x" : list(x_list),
                "y" : list(y_list),
            },
            "error" : self.error,
        }

        return format


def parse_excel(file: IO[bytes]) -> List[CoordinateMapping]:
    workbook = openpyxl.load_workbook(file)
    result = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))

        # Проверяем, что хотя бы есть заголовки
        if len(rows) < 2:
            result.append(
                CoordinateMapping(sheet_name=sheet_name,
                                  x_name="", y_name="",
                                  mapping=[], error=ErrorType.MISSING_HEADERS.value)
            )
            continue

        # Получение названий координат
        x_name, y_name = rows[0][:2]

        mapping = []


        for row in rows[1:]:
            if row[0] is None or row[1] is None: # Пропускаем пустые строки
                continue

            x_value = row[0]
            y_value = row[1]
            mapping.append((x_value, y_value))

        # Добавление результата в список
        result.append(CoordinateMapping(sheet_name=sheet_name, x_name=x_name, y_name=y_name, mapping=mapping).to_json())

    return result